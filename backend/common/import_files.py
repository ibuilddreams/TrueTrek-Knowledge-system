import csv
import io
import re

from openpyxl import Workbook, load_workbook


ALLOWED_EXTENSIONS = {".csv", ".xlsx"}
MAX_UPLOAD_SIZE_MB = 5


class ImportFileError(Exception):
    pass


def get_file_extension(filename):
    if not filename or "." not in filename:
        return ""
    return f".{filename.rsplit('.', 1)[-1].lower()}"


def validate_upload_file(uploaded_file):
    if uploaded_file is None:
        raise ImportFileError("A file is required.")

    extension = get_file_extension(uploaded_file.name)
    if extension not in ALLOWED_EXTENSIONS:
        raise ImportFileError("Only CSV and XLSX files are supported.")

    size = getattr(uploaded_file, "size", None)
    if size is not None and size > MAX_UPLOAD_SIZE_MB * 1024 * 1024:
        raise ImportFileError(f"File size must not exceed {MAX_UPLOAD_SIZE_MB}MB.")

    return extension


def _normalize_header(value):
    text = str(value or "").replace("\ufeff", "").replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _header_key(value):
    return re.sub(r"[^a-z0-9]+", "", _normalize_header(value).lower())


HEADER_ALIASES = {
    "firstname": "First Name",
    "lastname": "Last Name",
    "email": "Email",
    "password": "Password",
    "phone": "Phone",
    "phonenumber": "Phone",
    "gender": "Gender",
    "studentemail": "Student Email",
    "coursecode": "Course Code",
    "code": "Course Code",
    "coursetitle": "Course Title",
    "title": "Course Title",
}


def _canonicalize_header(value, required_headers, optional_headers=None):
    normalized = _normalize_header(value)
    known_headers = list(required_headers) + list(optional_headers or [])
    known_by_key = {_header_key(header): header for header in known_headers}

    key = _header_key(normalized)
    if key in known_by_key:
        return known_by_key[key]
    if key in HEADER_ALIASES and HEADER_ALIASES[key] in known_by_key.values():
        return HEADER_ALIASES[key]
    return normalized


def _row_to_dict(headers, values):
    row = {}
    for index, header in enumerate(headers):
        raw = values[index] if index < len(values) else ""
        if raw is None:
            raw = ""
        row[header] = str(raw).strip()
    return row


def parse_tabular_file(uploaded_file, required_headers, optional_headers=None, require_one_of=None):
    extension = validate_upload_file(uploaded_file)
    required = [_normalize_header(header) for header in required_headers]
    optional = [_normalize_header(header) for header in (optional_headers or [])]
    one_of = [_normalize_header(header) for header in (require_one_of or [])]

    if extension == ".csv":
        rows = _parse_csv(uploaded_file, required, optional, one_of)
    else:
        rows = _parse_xlsx(uploaded_file, required, optional, one_of)

    if not rows:
        raise ImportFileError("The file has no data rows.")

    return rows


def _parse_csv(uploaded_file, required_headers, optional_headers=None, require_one_of=None):
    uploaded_file.seek(0)
    raw = uploaded_file.read()
    if isinstance(raw, bytes):
        text = raw.decode("utf-8-sig")
    else:
        text = str(raw)

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    try:
        header_row = next(reader)
    except StopIteration as exc:
        raise ImportFileError("The file is empty.") from exc

    headers = [
        _canonicalize_header(cell, required_headers, optional_headers)
        for cell in header_row
    ]
    _validate_headers(headers, required_headers, require_one_of)

    rows = []
    for index, values in enumerate(reader, start=2):
        if not any(str(value or "").strip() for value in values):
            continue
        rows.append({"row_number": index, "data": _row_to_dict(headers, values)})
    return rows


def _parse_xlsx(uploaded_file, required_headers, optional_headers=None, require_one_of=None):
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)

    try:
        header_row = next(iterator)
    except StopIteration as exc:
        raise ImportFileError("The file is empty.") from exc

    headers = [
        _canonicalize_header(cell, required_headers, optional_headers)
        for cell in header_row
    ]
    _validate_headers(headers, required_headers, require_one_of)

    rows = []
    for index, values in enumerate(iterator, start=2):
        values = list(values or [])
        if not any(value is not None and str(value).strip() for value in values):
            continue
        rows.append({"row_number": index, "data": _row_to_dict(headers, values)})
    return rows


def _validate_headers(headers, required_headers, require_one_of=None):
    present = {_header_key(header) for header in headers}
    missing = [header for header in required_headers if _header_key(header) not in present]
    if missing:
        raise ImportFileError(
            f"Missing required column(s): {', '.join(missing)}. "
            f"Expected headers: {', '.join(required_headers)}."
        )

    if require_one_of:
        if not any(_header_key(header) in present for header in require_one_of):
            raise ImportFileError(
                "Missing course column. Provide either "
                f"{' or '.join(require_one_of)}."
            )


def build_sample_workbook(headers, sample_rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Template"
    sheet.append(headers)
    for row in sample_rows:
        sheet.append([row.get(header, "") for header in headers])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_sample_csv(headers, sample_rows):
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=headers)
    writer.writeheader()
    for row in sample_rows:
        writer.writerow({header: row.get(header, "") for header in headers})
    return io.BytesIO(buffer.getvalue().encode("utf-8-sig"))
